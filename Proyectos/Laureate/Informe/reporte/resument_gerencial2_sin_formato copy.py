import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill, Font

# Ruta del archivo actualizado
file_path_new = r'C:\Users\manue\OneDrive - EduCorpPERU\UPC - Calidad de Software\Proyectos\Pruebas Integrales\Planificación de Pruebas Integrales GL1 y GL3 v3.xlsx'

# Cargar el archivo Excel y la hoja correspondiente
plan_pruebas_integrales_df_new = pd.read_excel(file_path_new, sheet_name='Plan Pruebas Integrales')

# Convertir las columnas de fechas a formato de fecha, ignorando errores
plan_pruebas_integrales_df_new['Fecha Inicial'] = pd.to_datetime(plan_pruebas_integrales_df_new['Fecha Inicial'], format='%d/%m/%Y', errors='coerce')
plan_pruebas_integrales_df_new['Fecha Planificada'] = pd.to_datetime(plan_pruebas_integrales_df_new['Fecha Planificada'], format='%d/%m/%Y', errors='coerce')

# Función para contar los casos no cancelados y en un rango de fechas, por GL
def calcular_avance_planificado(df, gl_value, fecha_inicio, fecha_fin):
    df_filtrado = df[
        (df['N° GL'] == gl_value) & 
        (df['Estado Entregable'] != 'Cancelado') & 
        (df['Fecha Inicial'] >= fecha_inicio) & 
        (df['Fecha Inicial'] <= fecha_fin)
    ]
    casos_planificados = df_filtrado['N° CP'].nunique()
    return casos_planificados

# Definir la fecha de inicio y fecha de corte actual
fecha_inicio_antigua = plan_pruebas_integrales_df_new['Fecha Inicial'].min()
fecha_fin_actual = pd.to_datetime('24/10/2024', format='%d/%m/%Y')
fecha_corte_sem_ant = fecha_fin_actual - pd.Timedelta(days=7)

# Calcular el avance planificado para GL1 y GL3
avance_planificado_gl1_antigua = calcular_avance_planificado(plan_pruebas_integrales_df_new, 1, fecha_inicio_antigua, fecha_fin_actual)
avance_planificado_gl3_antigua = calcular_avance_planificado(plan_pruebas_integrales_df_new, 3, fecha_inicio_antigua, fecha_fin_actual)

# Función para contar los casos ejecutados hasta la fecha de corte
def contar_casos_ejecutados(df, gl_value, fecha_limite):
    df_filtrado = df[(df['N° GL'] == gl_value) & (df['Fecha Planificada'] <= fecha_limite)]
    casos_ejecutados = df_filtrado.groupby('N° CP').filter(
        lambda x: ((x['Estado Entregable'] == 'Aprobado') | (x['Estado Entregable'] == 'Pendiente retest')).any()
    )['N° CP'].nunique()
    return casos_ejecutados

# Calcular los casos ejecutados hasta la fecha de corte actual
ejecutados_gl1_new = contar_casos_ejecutados(plan_pruebas_integrales_df_new, 1, fecha_fin_actual)
ejecutados_gl3_new = contar_casos_ejecutados(plan_pruebas_integrales_df_new, 3, fecha_fin_actual)

# Calcular los ejecutados una semana antes (Real Sem. Ant.)
ejecutados_gl1_sem_ant = contar_casos_ejecutados(plan_pruebas_integrales_df_new, 1, fecha_corte_sem_ant)
ejecutados_gl3_sem_ant = contar_casos_ejecutados(plan_pruebas_integrales_df_new, 3, fecha_corte_sem_ant)

# Calcular los totales sumando los valores de GL1 y GL3
total_cant_pruebas = plan_pruebas_integrales_df_new[plan_pruebas_integrales_df_new['N° GL'] == 1]['N° CP'].nunique() + \
                     plan_pruebas_integrales_df_new[plan_pruebas_integrales_df_new['N° GL'] == 3]['N° CP'].nunique()

total_avance_planificado = avance_planificado_gl1_antigua + avance_planificado_gl3_antigua
total_avance_real = ejecutados_gl1_new + ejecutados_gl3_new
total_real_sem_ant = ejecutados_gl1_sem_ant + ejecutados_gl3_sem_ant

# Calcular el porcentaje respecto a Cant. Pruebas total
def calcular_porcentaje(valor, total):
    return f"{(valor / total * 100):.2f}%" if total > 0 else "0%"

# Calcular la variación numérica entre avance real y planificado
def calcular_variacion_numerica(real, planificado):
    return real - planificado

# Calcular la variación porcentual entre avance real y planificado
def calcular_variacion_porcentual(real, planificado):
    if planificado > 0:
        return f"{((real - planificado) / planificado * 100):.2f}%"
    else:
        return "0%"

# Crear la tabla de resultados con el porcentaje en la celda inferior
tabla_resultados_actualizada = pd.DataFrame({
    'N° GL': ['GL1', '', 'GL3', '', 'Total', ''],
    'Cant. Pruebas': [plan_pruebas_integrales_df_new[plan_pruebas_integrales_df_new['N° GL'] == 1]['N° CP'].nunique(), '100%',
                      plan_pruebas_integrales_df_new[plan_pruebas_integrales_df_new['N° GL'] == 3]['N° CP'].nunique(), '100%',
                      total_cant_pruebas, '100%'],
    'Avance Planificado': [avance_planificado_gl1_antigua, calcular_porcentaje(avance_planificado_gl1_antigua, total_cant_pruebas),
                           avance_planificado_gl3_antigua, calcular_porcentaje(avance_planificado_gl3_antigua, total_cant_pruebas),
                           total_avance_planificado, calcular_porcentaje(total_avance_planificado, total_cant_pruebas)],
    f'Avance Real {fecha_fin_actual.strftime("%d/%m")}': [ejecutados_gl1_new, calcular_porcentaje(ejecutados_gl1_new, total_cant_pruebas),
                                                          ejecutados_gl3_new, calcular_porcentaje(ejecutados_gl3_new, total_cant_pruebas),
                                                          total_avance_real, calcular_porcentaje(total_avance_real, total_cant_pruebas)],
    f'Real Sem. Ant. {fecha_corte_sem_ant.strftime("%d/%m")}': [ejecutados_gl1_sem_ant, calcular_porcentaje(ejecutados_gl1_sem_ant, total_cant_pruebas),
                                                                ejecutados_gl3_sem_ant, calcular_porcentaje(ejecutados_gl3_sem_ant, total_cant_pruebas),
                                                                total_real_sem_ant, calcular_porcentaje(total_real_sem_ant, total_cant_pruebas)],
    'Real vs Planificado (Var)': [calcular_variacion_numerica(ejecutados_gl1_new, avance_planificado_gl1_antigua), calcular_variacion_porcentual(ejecutados_gl1_new, avance_planificado_gl1_antigua),
                                  calcular_variacion_numerica(ejecutados_gl3_new, avance_planificado_gl3_antigua), calcular_variacion_porcentual(ejecutados_gl3_new, avance_planificado_gl3_antigua),
                                  calcular_variacion_numerica(total_avance_real, total_avance_planificado), calcular_variacion_porcentual(total_avance_real, total_avance_planificado)]
})

# Definir la ruta de salida para el archivo Excel
output_path = os.path.join(os.path.dirname(file_path_new), 'resultado_avance_pruebas_formateado.xlsx')

# Exportar los resultados a Excel
tabla_resultados_actualizada.to_excel(output_path, index=False)

# Cargar el archivo Excel para aplicar formato
wb = load_workbook(output_path)
ws = wb.active

# Aplicar formato de encabezado
fill_header = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
for cell in ws[1]:
    cell.fill = fill_header
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.font = Font(bold=True)

# Aplicar formato para centrar el contenido y poner color a la fila de totales
fill_total = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid")
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    for cell in row:
        cell.alignment = Alignment(horizontal='center', vertical='center')
        if cell.row == ws.max_row:
            cell.fill = fill_total
            cell.font = Font(bold=True)

# Guardar los cambios
wb.save(output_path)

print(f"El archivo ha sido guardado y formateado en: {output_path}")
