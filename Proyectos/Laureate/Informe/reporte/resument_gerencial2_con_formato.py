import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill, Font

# Ruta del archivo actualizado
file_path_new = r'C:\Users\manue\OneDrive - EduCorpPERU\UPC - Calidad de Software\Proyectos\Pruebas Integrales\Planificación de Pruebas Integrales GL1 y GL3 v3.xlsx'

# Cargar el archivo Excel y la hoja correspondiente
plan_pruebas_integrales_df_new = pd.read_excel(file_path_new, sheet_name='Plan Pruebas Integrales')

# Convertir las columnas de fechas a formato de fecha, ignorando errores
plan_pruebas_integrales_df_new['Fecha Inicial'] = pd.to_datetime(plan_pruebas_integrales_df_new['Fecha Inicial'], format='%d/%m/%Y', errors='coerce')
plan_pruebas_integrales_df_new['Fecha Planificada'] = pd.to_datetime(plan_pruebas_integrales_df_new['Fecha Planificada'], format='%d/%m/%Y', errors='coerce')

# Función para contar los casos aprobados
def contar_casos_aprobados(df, gl_value):
    df_filtrado = df[df['N° GL'] == gl_value]
    casos_aprobados = df_filtrado.groupby('N° CP').filter(
        lambda x: all(x['Estado Entregable'] == 'Aprobado')
    )['N° CP'].nunique()
    return casos_aprobados

# Definir la fecha de corte actual y la fecha más antigua
fecha_inicio_antigua = plan_pruebas_integrales_df_new['Fecha Inicial'].min()
fecha_fin_actual = pd.to_datetime('18/10/2024', format='%d/%m/%Y')

# Calcular los casos aprobados para GL1 y GL3
total_aprobados_gl1 = contar_casos_aprobados(plan_pruebas_integrales_df_new, 1)
total_aprobados_gl3 = contar_casos_aprobados(plan_pruebas_integrales_df_new, 3)
total_aprobados = total_aprobados_gl1 + total_aprobados_gl3

# Calcular otros datos para completar la tabla
total_cant_pruebas = plan_pruebas_integrales_df_new[plan_pruebas_integrales_df_new['N° GL'] == 1]['N° CP'].nunique() + \
                     plan_pruebas_integrales_df_new[plan_pruebas_integrales_df_new['N° GL'] == 3]['N° CP'].nunique()

avance_real_gl1 = 66  # Estos valores deben venir del cálculo de avance real
avance_real_gl3 = 12
avance_sem_ant_gl1 = 60
avance_sem_ant_gl3 = 7
real_vs_planificado_gl1 = -12
real_vs_planificado_gl3 = -58

# Crear la tabla de resultados incluyendo la columna "Total Aprobados"
tabla_resultados_actualizada = pd.DataFrame({
    'N° GL': ['GL1', 'GL3', 'Total'],
    'Cant. Pruebas': [plan_pruebas_integrales_df_new[plan_pruebas_integrales_df_new['N° GL'] == 1]['N° CP'].nunique(),
                      plan_pruebas_integrales_df_new[plan_pruebas_integrales_df_new['N° GL'] == 3]['N° CP'].nunique(),
                      total_cant_pruebas],
    'Avance Planificado': [78, 70, 148],  # Ejemplo de valores planificados
    'Avance Real 18/10': [avance_real_gl1, avance_real_gl3, avance_real_gl1 + avance_real_gl3],
    'Real Sem. Ant. 11/10': [avance_sem_ant_gl1, avance_sem_ant_gl3, avance_sem_ant_gl1 + avance_sem_ant_gl3],
    'Total Aprobados': [total_aprobados_gl1, total_aprobados_gl3, total_aprobados],
    'Real vs Planificado (Var)': [real_vs_planificado_gl1, real_vs_planificado_gl3, real_vs_planificado_gl1 + real_vs_planificado_gl3]
})

# Definir la ruta de salida para el archivo Excel
output_path = os.path.join(os.path.dirname(file_path_new), 'resultado_avance_pruebas_con_total_aprobados.xlsx')

# Exportar los resultados a Excel sin índice
tabla_resultados_actualizada.to_excel(output_path, index=False)

# Cargar el archivo Excel exportado para aplicar formato
wb = load_workbook(output_path)
ws = wb.active

# Aplicar formato de encabezado
fill_header = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
for cell in ws[1]:
    cell.fill = fill_header
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.font = Font(bold=True)

# Aplicar formato de filas con datos (centrar contenido y aplicar fondo)
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    for cell in row:
        cell.alignment = Alignment(horizontal='center', vertical='center')

# Aplicar formato de la fila de totales
for cell in ws[ws.max_row]:
    cell.font = Font(bold=True)

# Guardar los cambios
wb.save(output_path)

print(f"El archivo ha sido guardado y formateado en: {output_path}")