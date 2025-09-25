import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle
from openpyxl.styles.numbers import FORMAT_DATE_XLSX15

# Load the specific sheet and specify columns of interest
file_path = r'C:\Users\manue\OneDrive - EduCorpPERU\UPC - Calidad de Software\Proyectos\Pruebas Integrales\Planificación de Pruebas Integrales GL1 y GL3 v3.xlsx'
columns_of_interest = ['N° GL', 'N° CP', 'Caso de Prueba','Fecha Replanificada CP', 'BPO', 'PM Asignado', 'Periférico', 'Hora de Inicio', 'Hora de Fin']
data = pd.read_excel(file_path, sheet_name='Plan Pruebas Integrales', usecols=columns_of_interest)

# Ensure date columns are datetime type
data['Fecha Replanificada CP'] = pd.to_datetime(data['Fecha Replanificada CP'], errors='coerce')

# Define the filter function with optional multiple "N° GL" and "Estado CP" values
def filter_data(df, start_date, end_date, gl_numbers=None, estados_cp=None):
    start_date = pd.to_datetime(start_date)
    end_date = pd.to_datetime(end_date)
    filtered_df = df[(df['Fecha Replanificada CP'] >= start_date) & 
                     (df['Fecha Replanificada CP'] <= end_date)]
    
    # Apply optional filter for multiple "N° GL" values if provided
    if gl_numbers:
        filtered_df = filtered_df[filtered_df['N° GL'].isin(gl_numbers)]
    
    # Apply optional filter for multiple "Estado CP" values if provided
    if estados_cp:
        filtered_df = filtered_df[filtered_df['Estado CP'].isin(estados_cp)]
    
    return filtered_df

# Example usage
start_date = '2024-11-11'
end_date = '2024-11-15'
gl_numbers = [1, 3]  # Optional: List of "N° GL" values to filter by, e.g., [1, 3]
estados_cp = ['Pendiente retest', 'Planificado']  # Optional: List of states to filter by
filtered_data = filter_data(data, start_date, end_date, gl_numbers, estados_cp)

# Add 'Tipo CP' and 'Caso de Prueba' columns (default empty or add logic to populate if necessary)
filtered_data['Tipo CP'] = ''  # Fill with appropriate logic if available
filtered_data['Caso de Prueba'] = ''  # Fill with appropriate logic if available

# Reorder the columns as specified
ordered_columns = ['N° GL', 'N° CP', 'Tipo CP', 'Caso de Prueba', 'Fecha Replanificada CP', 'Hora de Inicio', 'Hora de Fin', 'Periférico', 'PM Asignado', 'BPO']
filtered_data = filtered_data[ordered_columns]

# Define the output file path
output_file_name = f'reporte_{start_date}_a_{end_date}.xlsx'
output_file_path = r'G:\My Drive\Data Analysis\PyTools\Test Integrales\Informe\reporte_' + output_file_name

# Export the filtered data to an Excel file without formatting (temporary step)
filtered_data.to_excel(output_file_path, index=False)

# Open the Excel file to apply formatting
wb = load_workbook(output_file_path)
ws = wb.active

# Create a named style for date format (short date)
date_style = NamedStyle(name="date_style", number_format=FORMAT_DATE_XLSX15)

# Apply the date format to the 'Fecha Replanificada CP' column (which is the 5th column, index starts at 1)
for row in ws.iter_rows(min_row=2, min_col=5, max_col=5, max_row=ws.max_row):
    for cell in row:
        cell.style = date_style

# Save the workbook with the formatted date
wb.save(output_file_path)

print(f"\nFiltered data with formatted 'Fecha Replanificada CP' has been exported to: {output_file_path}")
