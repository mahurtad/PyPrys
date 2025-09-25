import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Ruta del archivo de origen
file_path = r'C:\Users\manue\OneDrive - EduCorpPERU\UPC - Calidad de Software\Proyectos\Pruebas Integrales\Planificación de Pruebas Integrales GL1 y GL3 v3.xlsx'

# Ruta de salida del archivo consolidado
output_path = r'C:\Users\manue\OneDrive - EduCorpPERU\UPC - Calidad de Software\Proyectos\Pruebas Integrales\planificados_resaltados.xlsx'

# Rango de cortes de fechas
date_ranges = [
    ('2024-11-25', '2024-11-30'),
    ('2024-12-02', '2024-12-08'),
    ('2024-12-09', '2024-12-15'),
    ('2024-12-16', '2024-12-22'),
    ('2024-12-23', '2024-12-29'),
    ('2024-12-30', '2025-01-05'),
    ('2025-01-06', '2025-01-12'),
    ('2025-01-13', '2025-01-19'),
    ('2025-01-20', '2025-01-26'),
    ('2025-01-27', '2025-02-02'),
    ('2025-02-03', '2025-02-09'),
    ('2025-02-10', '2025-02-16'),
    ('2025-02-17', '2025-02-23'),
    ('2025-02-24', '2025-03-02')
]

# Cargar los datos desde la hoja "Plan Pruebas Integrales"
data = pd.read_excel(file_path, sheet_name='Plan Pruebas Integrales')

# Convertir la columna "Fecha Planificada" a formato datetime
data['Fecha Planificada'] = pd.to_datetime(data['Fecha Planificada'], errors='coerce')

# Definir las columnas de interés
columns_of_interest = [
    'N° GL', 'N° CP', 'N° D', 'Caso de Prueba',
    'Nombre de Tarea (Desarrollo)', 'Fecha Planificada', 'BPO', 'PM Asignado', 'Periférico', 'Estado Entregable'
]

# Crear el archivo Excel consolidado
with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
    for start_date, end_date in date_ranges:
        # Filtrar los datos para el rango de fechas actual
        filtered_data = data[
            (data['Fecha Planificada'] >= pd.to_datetime(start_date)) &
            (data['Fecha Planificada'] <= pd.to_datetime(end_date)) &
            (~data['Estado Entregable'].isin(['Aprobado', 'Pendiente aprobación']))
        ][columns_of_interest]

        # Identificar el registro con la fecha más próxima para cada caso de prueba
        if not filtered_data.empty:
            closest_date = filtered_data.loc[
                filtered_data.groupby('N° CP')['Fecha Planificada'].idxmin()
            ]
            # Agregar una columna temporal para marcar las filas con la fecha más próxima
            filtered_data['Resaltar'] = filtered_data.index.isin(closest_date.index)

        # Generar un nombre para la hoja basado en el rango de fechas
        sheet_name = f'{start_date}_a_{end_date}'

        # Escribir los datos filtrados en la hoja correspondiente
        filtered_data.to_excel(writer, sheet_name=sheet_name, index=False)

# Resaltar las celdas correspondientes en el archivo Excel
wb = load_workbook(output_path)
highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Color amarillo

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    # Aplicar el formato de fecha corta DD/MM/AAAA a la columna Fecha Planificada
    for row in ws.iter_rows(min_row=2, min_col=6, max_col=6, max_row=ws.max_row):  # Columna de Fecha Planificada
        for cell in row:
            cell.number_format = "DD/MM/YYYY"
    # Resaltar las filas con 'Resaltar' marcado como True
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):  # Excluir encabezados
        if row[-1].value:  # La última columna es 'Resaltar'
            for cell in row:
                cell.fill = highlight_fill

    # Eliminar la columna 'Resaltar' del archivo final
    ws.delete_cols(len(columns_of_interest) + 1)  # Última columna

# Guardar el libro de trabajo con el formato aplicado
wb.save(output_path)

print(f"Archivo consolidado con resaltados creado exitosamente en: {output_path}")
