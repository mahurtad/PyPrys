import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill, Font

# Ruta del archivo actualizado
file_path_new = r'C:\Users\manue\OneDrive - EduCorpPERU\UPC - Calidad de Software\Proyectos\Pruebas Integrales\Planificación de Pruebas Integrales GL1 y GL3 v3.xlsx'

# Cargar el archivo Excel y la hoja correspondiente
plan_pruebas_integrales_df_new = pd.read_excel(file_path_new, sheet_name='Plan Pruebas Integrales')

# Convertir las columnas de fechas a formato de fecha, ignorando errores
plan_pruebas_integrales_df_new['Fecha Inicial'] = pd.to_datetime(plan_pruebas_integrales_df_new['Fecha Inicial'], format='%d/%m/%Y', errors='coerce')
plan_pruebas_integrales_df_new['Fecha Planificada'] = pd.to_datetime(plan_pruebas_integrales_df_new['Fecha Planificada'], format='%d/%m/%Y', errors='coerce')

# Función para contar los casos no cancelados y en un rango de fechas, por GL (avance planificado)
def calcular_avance_planificado(df, gl_value, fecha_inicio, fecha_fin):
    df_filtrado = df[(df['N° GL'] == gl_value) & (df['Estado Entregable'] != 'Cancelado') & 
                     (df['Fecha Inicial'] >= fecha_inicio) & (df['Fecha Inicial'] <= fecha_fin)]
    return df_filtrado['N° CP'].nunique()

# Función para contar los casos ejecutados hasta la fecha de corte (avance real)
def contar_casos_ejecutados(df, gl_value, fecha_limite):
    df_filtrado = df[(df['N° GL'] == gl_value) & (df['Fecha Planificada'] <= fecha_limite)]
    return df_filtrado.groupby('N° CP').filter(lambda x: ((x['Estado Entregable'] == 'Aprobado') | 
                                                          (x['Estado Entregable'] == 'Pendiente retest')).any())['N° CP'].nunique()

# Función para contar los casos aprobados donde todos los registros de "N° CP" están aprobados
def contar_casos_aprobados(df, gl_value):
    # Filtrar solo los casos del GL especificado
    df_filtrado = df[df['N° GL'] == gl_value]
    # Agrupar por "N° CP" y verificar que todos los registros de cada caso están en estado "Aprobado"
    aprobados_df = df_filtrado.groupby('N° CP').filter(lambda x: (x['Estado Entregable'] == 'Aprobado').all())
    return aprobados_df['N° CP'].nunique()

# Definir la fecha de inicio y la fecha de corte actual
fecha_inicio_antigua = plan_pruebas_integrales_df_new['Fecha Inicial'].min()
fecha_fin_actual = pd.to_datetime('21/11/2024', format='%d/%m/%Y')
fecha_corte_sem_ant = fecha_fin_actual - pd.Timedelta(days=7)

# Calcular dinámicamente los datos
avance_planificado_gl1 = calcular_avance_planificado(plan_pruebas_integrales_df_new, 1, fecha_inicio_antigua, fecha_fin_actual)
avance_planificado_gl3 = calcular_avance_planificado(plan_pruebas_integrales_df_new, 3, fecha_inicio_antigua, fecha_fin_actual)

avance_real_gl1 = contar_casos_ejecutados(plan_pruebas_integrales_df_new, 1, fecha_fin_actual)
avance_real_gl3 = contar_casos_ejecutados(plan_pruebas_integrales_df_new, 3, fecha_fin_actual)

avance_real_sem_ant_gl1 = contar_casos_ejecutados(plan_pruebas_integrales_df_new, 1, fecha_corte_sem_ant)
avance_real_sem_ant_gl3 = contar_casos_ejecutados(plan_pruebas_integrales_df_new, 3, fecha_corte_sem_ant)

aprobados_gl1 = contar_casos_aprobados(plan_pruebas_integrales_df_new, 1)
aprobados_gl3 = contar_casos_aprobados(plan_pruebas_integrales_df_new, 3)

# Calcular la variación real vs planificado
var_real_vs_plan_gl1 = avance_real_gl1 - avance_planificado_gl1
var_real_vs_plan_gl3 = avance_real_gl3 - avance_planificado_gl3

# Calcular el porcentaje de aprobados sobre real y sobre total
def calcular_porcentaje_aprobado_sobre_real(aprobados, avance_real):
    return f"{(aprobados / avance_real * 100):.2f}%" if avance_real > 0 else "0%"

def calcular_porcentaje_aprobado_sobre_total(aprobados, cant_pruebas):
    return f"{(aprobados / cant_pruebas * 100):.2f}%" if cant_pruebas > 0 else "0%"

# Calcular dinámicamente el total de casos de prueba por GL
cant_pruebas_gl1 = plan_pruebas_integrales_df_new[plan_pruebas_integrales_df_new['N° GL'] == 1]['N° CP'].nunique()
cant_pruebas_gl3 = plan_pruebas_integrales_df_new[plan_pruebas_integrales_df_new['N° GL'] == 3]['N° CP'].nunique()

# Crear la tabla Cuadro Avance con todos los cálculos dinámicos
tabla_resultados_actualizada = pd.DataFrame({
    'N° GL': ['GL1', 'GL3', 'Total'],
    'Cant. Pruebas': [cant_pruebas_gl1, cant_pruebas_gl3, cant_pruebas_gl1 + cant_pruebas_gl3],
    'Avance Planificado': [avance_planificado_gl1, avance_planificado_gl3, avance_planificado_gl1 + avance_planificado_gl3],
    'Avance Real 24/10': [avance_real_gl1, avance_real_gl3, avance_real_gl1 + avance_real_gl3],
    'Real Sem. Ant. 17/10': [avance_real_sem_ant_gl1, avance_real_sem_ant_gl3, avance_real_sem_ant_gl1 + avance_real_sem_ant_gl3],
    'Real vs Planificado (Var)': [var_real_vs_plan_gl1, var_real_vs_plan_gl3, var_real_vs_plan_gl1 + var_real_vs_plan_gl3],
    'Aprobados': [aprobados_gl1, aprobados_gl3, aprobados_gl1 + aprobados_gl3],
    '% Aprobado (sobre Real)': [calcular_porcentaje_aprobado_sobre_real(aprobados_gl1, avance_real_gl1),
                                calcular_porcentaje_aprobado_sobre_real(aprobados_gl3, avance_real_gl3),
                                calcular_porcentaje_aprobado_sobre_real(aprobados_gl1 + aprobados_gl3, avance_real_gl1 + avance_real_gl3)],
    '% Aprobado (sobre Total)': [calcular_porcentaje_aprobado_sobre_total(aprobados_gl1, cant_pruebas_gl1),
                                 calcular_porcentaje_aprobado_sobre_total(aprobados_gl3, cant_pruebas_gl3),
                                 calcular_porcentaje_aprobado_sobre_total(aprobados_gl1 + aprobados_gl3, cant_pruebas_gl1 + cant_pruebas_gl3)]
})

# Exportar los resultados a Excel
output_path = os.path.join(os.path.dirname(file_path_new), 'resumen_gerencial_v3_part2.xlsx')

# Guardar la tabla en Excel con formato
with pd.ExcelWriter(output_path, engine='openpyxl', mode='w') as writer:
    tabla_resultados_actualizada.to_excel(writer, sheet_name='Cuadro Avance', index=False)

# Cargar el archivo Excel para aplicar formato
wb = load_workbook(output_path)
ws_avance = wb['Cuadro Avance']

# Aplicar formato de encabezado
fill_header = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
fill_total = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid")

# Centrar todo el contenido y aplicar color
for row in ws_avance.iter_rows(min_row=1, max_row=ws_avance.max_row, min_col=1, max_col=ws_avance.max_column):
    for cell in row:
        cell.alignment = Alignment(horizontal='center', vertical='center')
        if cell.row == 1:
            cell.fill = fill_header
            cell.font = Font(bold=True)
        elif cell.row == ws_avance.max_row:
            cell.fill = fill_total
            cell.font = Font(bold=True)

# Guardar los cambios
wb.save(output_path)

print(f"El archivo ha sido guardado en: {output_path}")
