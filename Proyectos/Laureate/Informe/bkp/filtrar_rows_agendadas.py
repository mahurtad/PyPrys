import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle
from openpyxl.styles.numbers import FORMAT_DATE_XLSX15

# Load the specific sheet "Programación Pruebas" and specify all columns of interest
file_path = r'C:\Users\manue\OneDrive - EduCorpPERU\UPC - Calidad de Software\Proyectos\Pruebas Integrales\Planificación de Pruebas Integrales GL1 y GL3 v3.xlsx'
columns_of_interest = ['Tipo CP', 'Nombre Prueba', 'Fecha', 'Hora de Inicio', 'Hora Fin', 'Periférico', 'PM', 'BPO', 
                       'QA Lider Prueba', 'QA Crea Informe', 'ESTADO', 'Comentarios', 'Link de grabación de la sesión']
data = pd.read_excel(file_path, sheet_name='Programación Pruebas', usecols=columns_of_interest)

# Ensure date columns are datetime type
data['Fecha'] = pd.to_datetime(data['Fecha'], errors='coerce')

# Define the filter function to filter based on "Fecha"
def filter_data_by_date(df, start_date, end_date):
    start_date = pd.to_datetime(start_date)
    end_date = pd.to_datetime(end_date)
    
    # Filter data based on the "Fecha" column
    filtered_df = df[(df['Fecha'] >= start_date) & 
                     (df['Fecha'] <= end_date)]
    
    return filtered_df

# Example usage
start_date = '2024-10-09'
end_date = '2024-10-09'

# Apply the filter based on the date range
filtered_data = filter_data_by_date(data, start_date, end_date)

# Display the filtered data in the console
print(filtered_data)

# Display the count of unique "Nombre Prueba" values in the filtered data
unique_test_count = filtered_data['Nombre Prueba'].nunique()
print(f"\nTotal unique 'Nombre Prueba' records: {unique_test_count}")

# Define the output file name based on the date range
output_file_name = f'planificadas_{start_date}_a_{end_date}.xlsx'
output_file_path = rf'G:\My Drive\Data Analysis\PyTools\Test Integrales\reportes\{output_file_name}'

# Export the filtered data to an Excel file without formatting (temporary step)
filtered_data.to_excel(output_file_path, index=False)

# Open the Excel file to apply formatting
wb = load_workbook(output_file_path)
ws = wb.active

# Create a named style for date format (short date)
date_style = NamedStyle(name="date_style", number_format=FORMAT_DATE_XLSX15)

# Apply the date format to the 'Fecha' column (which is the 3rd column, index starts at 1)
for row in ws.iter_rows(min_row=2, min_col=3, max_col=3, max_row=ws.max_row):
    for cell in row:
        cell.style = date_style

# Save the workbook with the formatted date
wb.save(output_file_path)

print(f"\nFiltered data with formatted 'Fecha' has been exported to: {output_file_path}")
