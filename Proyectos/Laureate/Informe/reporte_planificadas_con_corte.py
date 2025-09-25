import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle

# Ruta del archivo de origen
file_path = r'C:\Users\manue\OneDrive - EduCorpPERU\UPC - Calidad de Software\Proyectos\Pruebas Integrales\Planificación de Pruebas Integrales GL1 y GL3 v3.xlsx'

# Definir el rango de fechas para el reporte
start_date = '2024-11-28'
end_date = '2024-11-30'

# Crear el nombre del archivo de salida basado en el rango de fechas
output_path = f'C:\\Users\\manue\\OneDrive - EduCorpPERU\\UPC - Calidad de Software\\Proyectos\\Pruebas Integrales\\planificados_{start_date}_a_{end_date}.xlsx'

# Cargar los datos desde la hoja "Plan Pruebas Integrales"
data = pd.read_excel(file_path, sheet_name='Plan Pruebas Integrales')

# Convertir la columna "Fecha Planificada" a formato datetime
data['Fecha Planificada'] = pd.to_datetime(data['Fecha Planificada'], errors='coerce')

# Definir los campos necesarios para ambos reportes
columns_of_interest = [
    'N° GL', 'N° CP', 'N° D', 'Caso de Prueba',
    'Nombre de Tarea (Desarrollo)', 'Fecha Planificada', 'BPO', 'PM Asignado', 'Periférico', 'Estado Entregable'
]

# Filtrar los registros que están en el rango de fechas y con estado diferente de "Aprobado" o "Pendiente aprobación"
filtered_data = data[
    (data['Fecha Planificada'] >= pd.to_datetime(start_date)) &
    (data['Fecha Planificada'] <= pd.to_datetime(end_date)) &
    (~data['Estado Entregable'].isin(['Aprobado', 'Pendiente aprobación']))
][columns_of_interest]

# Crear un estilo de fecha corta
date_style = NamedStyle(name="date_style", number_format="DD/MM/YYYY")

# Guardar ambas hojas en el archivo y aplicar formato de fecha
with pd.ExcelWriter(output_path, engine='openpyxl', mode='w') as writer:
    filtered_data.to_excel(writer, sheet_name='Entregables x Semana', index=False)
    # Generar el segundo reporte basado en los datos filtrados en la primera hoja, agrupando los campos
    grouped_data = filtered_data.groupby(['N° GL', 'N° CP']).agg({
        'N° D': lambda x: ', '.join(x.dropna().astype(str).unique()),
        'Caso de Prueba': 'first',
        'Nombre de Tarea (Desarrollo)': lambda x: ', '.join(x.dropna().unique()),
        'Fecha Planificada': 'first',
        'BPO': lambda x: ', '.join(x.dropna().unique()),
        'PM Asignado': lambda x: ', '.join(x.dropna().unique()),
        'Periférico': lambda x: ', '.join(x.dropna().unique()),
        'Estado Entregable': 'first'
    }).reset_index()
    
    # Seleccionar solo las columnas necesarias para la segunda hoja
    grouped_data = grouped_data[columns_of_interest]
    grouped_data.to_excel(writer, sheet_name='Casos x Semana', index=False)

# Aplicar formato de fecha a "Fecha Planificada" en ambas hojas
wb = load_workbook(output_path)
for sheet_name in ['Entregables x Semana', 'Casos x Semana']:
    ws = wb[sheet_name]
    for row in ws.iter_rows(min_row=2, min_col=6, max_col=6, max_row=ws.max_row):
        for cell in row:
            cell.number_format = "DD/MM/YYYY"  # Aplicar formato de fecha corta

# Guardar el libro de trabajo con la fecha formateada
wb.save(output_path)

print(f"El archivo consolidado ha sido guardado en: {output_path}")
