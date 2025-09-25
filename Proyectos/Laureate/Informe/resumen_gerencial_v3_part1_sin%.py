import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill, Font

# Ruta del archivo actualizado
file_path_new = r'C:\Users\manue\OneDrive - EduCorpPERU\UPC - Calidad de Software\Proyectos\Pruebas Integrales\Planificación de Pruebas Integrales GL1 y GL3 v3.xlsx'

# Definir los datos que se deben mostrar en la hoja Cuadro Avance
avance_real_gl1 = 70
avance_real_gl3 = 22
avance_real_sem_ant_gl1 = 65
avance_real_sem_ant_gl3 = 10
avance_planificado_gl1 = 78
avance_planificado_gl3 = 75
var_real_vs_plan_gl1 = avance_real_gl1 - avance_planificado_gl1
var_real_vs_plan_gl3 = avance_real_gl3 - avance_planificado_gl3

# Agregar columna "Cant. Pruebas" con los respectivos valores
cant_pruebas_gl1 = 95
cant_pruebas_gl3 = 98

# Crear la tabla Cuadro Avance con los valores y agregar la columna "Cant. Pruebas"
tabla_resultados_actualizada = pd.DataFrame({
    'N° GL': ['GL1', 'GL3', 'Total'],
    'Cant. Pruebas': [cant_pruebas_gl1, cant_pruebas_gl3, cant_pruebas_gl1 + cant_pruebas_gl3],
    'Avance Planificado': [avance_planificado_gl1, avance_planificado_gl3, avance_planificado_gl1 + avance_planificado_gl3],
    'Avance Real 24/10': [avance_real_gl1, avance_real_gl3, avance_real_gl1 + avance_real_gl3],
    'Real Sem. Ant. 17/10': [avance_real_sem_ant_gl1, avance_real_sem_ant_gl3, avance_real_sem_ant_gl1 + avance_real_sem_ant_gl3],
    'Real vs Planificado (Var)': [var_real_vs_plan_gl1, var_real_vs_plan_gl3, var_real_vs_plan_gl1 + var_real_vs_plan_gl3]
})

# Exportar los resultados a Excel
output_path = os.path.join(os.path.dirname(file_path_new), 'resultado_avance_pruebas_con_cant_pruebas.xlsx')

# Guardar la tabla en Excel con formato
with pd.ExcelWriter(output_path, engine='openpyxl', mode='w') as writer:
    tabla_resultados_actualizada.to_excel(writer, sheet_name='Cuadro Avance', index=False)

# Cargar el archivo Excel para aplicar formato
wb = load_workbook(output_path)
ws_avance = wb['Cuadro Avance']

# Aplicar formato de encabezado
fill_header = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
fill_total = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid")

# Centrar todo el contenido y aplicar color
for row in ws_avance.iter_rows(min_row=1, max_row=ws_avance.max_row, min_col=1, max_col=ws_avance.max_column):
    for cell in row:
        cell.alignment = Alignment(horizontal='center', vertical='center')
        if cell.row == 1:
            cell.fill = fill_header
            cell.font = Font(bold=True)
        elif cell.row == ws_avance.max_row:
            cell.fill = fill_total
            cell.font = Font(bold=True)

# Guardar los cambios
wb.save(output_path)

print(f"El archivo ha sido guardado en: {output_path}")
