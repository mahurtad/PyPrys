import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import NamedStyle
from openpyxl.styles.numbers import FORMAT_DATE_XLSX15

# Load the specific sheet and specify columns of interest
file_path = r'C:\Users\manue\OneDrive - EduCorpPERU\UPC - Calidad de Software\Proyectos\Pruebas Integrales\Planificación de Pruebas Integrales GL1 y GL3 v3.xlsx'
columns_of_interest = ['N° GL', 'N° CP', 'N° D','Caso de Prueba', 'Fecha Replanificada CP', 'BPO', 'PM Asignado', 'Periférico', 'Estado CP', 'Hora de Inicio', 'Hora de Fin']
data = pd.read_excel(file_path, sheet_name='Plan Pruebas Integrales', usecols=columns_of_interest)

# Ensure date columns are datetime type
data['Fecha Replanificada CP'] = pd.to_datetime(data['Fecha Replanificada CP'], errors='coerce')

# Define the filter function with optional multiple "N° GL" and "Estado CP" values
def filter_data(df, start_date, end_date, gl_numbers=None, estados_cp=None):
    start_date = pd.to_datetime(start_date)
    end_date = pd.to_datetime(end_date)
    filtered_df = df[(df['Fecha Replanificada CP'] >= start_date) & 
                     (df['Fecha Replanificada CP'] <= end_date)]
    
    # Apply optional filter for multiple "N° GL" values if provided
    if gl_numbers:
        filtered_df = filtered_df[filtered_df['N° GL'].isin(gl_numbers)]
    
    # Apply optional filter for multiple "Estado CP" values if provided
    if estados_cp:
        filtered_df = filtered_df[filtered_df['Estado CP'].isin(estados_cp)]
    
    return filtered_df

# Example usage
start_date = '2024-11-11'
end_date = '2024-11-15'
gl_numbers = [1, 3]  # Optional: List of "N° GL" values to filter by, e.g., [1, 3]
estados_cp = ['Pendiente retest', 'Planificado','Replanificado']  # Optional: List of states to filter by
filtered_data = filter_data(data, start_date, end_date, gl_numbers, estados_cp)

# Display the filtered data
print(filtered_data)

# Display the count of unique "N° CP" values in the filtered data
unique_cp_count = filtered_data['N° CP'].nunique()
print(f"\nTotal unique 'N° CP' records: {unique_cp_count}")

# Reorder the columns as specified
ordered_columns = ['N° GL', 'N° CP','Tipo CP','Caso de Prueba','Fecha Replanificada CP','Hora de Inicio', 'Hora de Fin','Periférico','PM Asignado', 'BPO']

# Since 'Tipo CP' and 'Caso de Prueba' are not in the original data, let's assume you want to fill them with empty values or default data
filtered_data['Tipo CP'] = ''  # Fill with empty values or use a default
#filtered_data['Caso de Prueba'] = ''  # Fill with empty values or use a default

# Reorder the columns
filtered_data = filtered_data[ordered_columns]

# Define the output file path
output_file_path = r'G:\My Drive\Data Analysis\PyTools\Test Integrales\reportes\programación_diaria.xlsx'

# Export the filtered data to an Excel file without formatting (temporary step)
filtered_data.to_excel(output_file_path, index=False)

# Open the Excel file to apply formatting
wb = load_workbook(output_file_path)
ws = wb.active

# Create a named style for date format (short date)
date_style = NamedStyle(name="date_style", number_format=FORMAT_DATE_XLSX15)

# Apply the date format to the 'Fecha Replanificada CP' column (which is the 5th column, index starts at 1)
for row in ws.iter_rows(min_row=2, min_col=5, max_col=5, max_row=ws.max_row):
    for cell in row:
        cell.style = date_style

# Save the workbook with the formatted date
wb.save(output_file_path)

print(f"\nFiltered data with formatted 'Fecha Replanificada CP' has been exported to: {output_file_path}")
