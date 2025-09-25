from __future__ import annotations
import re
import shutil
from datetime import datetime
from typing import Optional
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook

# === CONFIGURACIÓN ===
EXCEL_PATH = Path("C:/Users/manue/OneDrive - EduCorpPERU/Calidad de Software - Certificaciones/Gestión Demanda Certificaciones LIUv1.xlsx")
SHEET_NAME = "Tickets Diarios"
DEST_ROOT = Path("C:/Users/manue/OneDrive - EduCorpPERU/2025/Tickets")

COL_RITM_INC      = "RITM/INC"
COL_SCTASK        = "SCTASK"
COL_APLICATIVO    = "APLICATIVO/BD"
COL_DESCRIPCION   = "DESCRIPCIÓN"
COL_DESARROLLADOR = "DESARROLLADOR"
ULTIMO_SEGMENTO   = "MAHURTAD"

# Rutas de plantilla
PLANTILLA_DIR = Path("C:/Users/manue/OneDrive - EduCorpPERU/2025/Tickets/Plantilla/16222023-RITMINCXXX-SISTEMA-TITULO-MAHURTAD")
PLANTILLA_INF = Path("C:/Users/manue/OneDrive - EduCorpPERU/2025/Tickets/Plantilla/INF-DDMMAAAA-RITMXXXXXXX-SCTASKXXXX-{SISTEMA}-{TITULO}.xlsx")
SUBCARPETA_INF = Path("3. Informe Certificación")

# Carpeta Descargas
DOWNLOADS_ROOT = Path("C:/Users/manue/Downloads")

# Opcional: recorte del 5to segmento (DESCRIPCIÓN) para evitar nombres larguísimos
MAX_DESC_LEN = 100
INVALID_CHARS_PATTERN = r'[\\/:*?"<>|]'

def log(msg: str) -> None:
    print(msg, flush=True)

def sanitize_segment(text: str) -> str:
    """Quita caracteres inválidos de Windows, colapsa espacios y recorta extremos."""
    if pd.isna(text):
        return ""
    clean = re.sub(INVALID_CHARS_PATTERN, " ", str(text))
    clean = re.sub(r"\s+", " ", clean, flags=re.UNICODE).strip()
    return clean

def shorten(text: str, max_len: int) -> str:
    if max_len and len(text) > max_len:
        return text[:max_len].rstrip()
    return text

def build_folder_name(ritm_inc: str, sctask: str, aplicativo: str, descripcion: str) -> str:
    """Construye: DDMMAAAA-RITM/INC-SCTASK-APLICATIVO/BD-DESCRIPCIÓN-MAHURTAD"""
    fecha = datetime.now().strftime("%d%m%Y")
    parts = [
        fecha,
        sanitize_segment(ritm_inc),
        sanitize_segment(sctask),
        sanitize_segment(aplicativo),
        shorten(sanitize_segment(descripcion), MAX_DESC_LEN),
        ULTIMO_SEGMENTO,
    ]
    parts = [p for p in parts if p]
    name = "-".join(parts)
    name = re.sub(r"\s+", " ", name).strip()
    return name

def find_ticket_row(df: pd.DataFrame, ritm_inc_value: str):
    target = str(ritm_inc_value).strip().lower()
    mask = df[COL_RITM_INC].astype(str).str.strip().str.lower() == target
    matches = df[mask]
    if matches.empty:
        return None
    return matches.iloc[0]

def count_items_in_dir(path: Path) -> tuple[int, int]:
    files, dirs = 0, 0
    for entry in path.iterdir():
        if entry.is_file():
            files += 1
        elif entry.is_dir():
            dirs += 1
    return files, dirs

def copy_template_into(target_folder: Path) -> bool:
    """Copia el contenido de PLANTILLA_DIR dentro de target_folder con logs."""
    if not PLANTILLA_DIR.is_dir():
        log(f"⚠️  Plantilla no encontrada:\n{PLANTILLA_DIR}")
        return False

    fcount, dcount = count_items_in_dir(PLANTILLA_DIR)
    log(f"Copiando plantilla… ({dcount} carpetas, {fcount} archivos)")

    copied = 0
    for src in PLANTILLA_DIR.iterdir():
        dst = target_folder / src.name
        try:
            if src.is_dir():
                shutil.copytree(src, dst, dirs_exist_ok=True)
                log(f"  📁 {src.name}  ->  {dst}")
            else:
                dst.parent.mkdir(parents=True, exist_ok=True)
                shutil.copy2(src, dst)
                log(f"  📄 {src.name}  ->  {dst}")
            copied += 1
        except Exception as e:
            log(f"  ⚠️  No se pudo copiar '{src}' -> '{dst}'. Detalle: {e}")
    log(f"Plantilla copiada. Elementos procesados: {copied}")
    return True

def copy_and_rename_inf(target_folder: Path, folder_name: str) -> Optional[Path]:
    """
    Copia PLANTILLA_INF a '3. Informe Certificación' y la renombra a:
      INF-<folder_name_sin_-MAHURTAD>.xlsx
    Devuelve la ruta final del archivo INF generado, o None si falla.
    """
    if not PLANTILLA_INF.is_file():
        log(f"⚠️  Archivo plantilla INF no encontrado:\n{PLANTILLA_INF}")
        return None

    base_without_last = folder_name
    suffix = f"-{ULTIMO_SEGMENTO}"
    if base_without_last.endswith(suffix):
        base_without_last = base_without_last[:-len(suffix)]

    new_filename = f"INF-{base_without_last}.xlsx"
    dest_subfolder = target_folder / SUBCARPETA_INF
    dest_subfolder.mkdir(parents=True, exist_ok=True)
    dest_path = dest_subfolder / new_filename

    try:
        shutil.copy2(PLANTILLA_INF, dest_path)
        log(f"INF copiado y renombrado:\n{dest_path}")
        return dest_path
    except Exception as e:
        log(f"❌ Error al copiar/renombrar INF:\nOrigen: {PLANTILLA_INF}\nDestino: {dest_path}\nDetalle: {e}")
        return None

def ensure_downloads_folder_and_copy(ritm_inc_input: str, inf_path: Optional[Path]) -> None:
    """
    Crea SIEMPRE C:/Users/manue/Downloads/<RITM_O_INC>.
    Si inf_path existe, copia el INF allí también.
    """
    folder_name = sanitize_segment(ritm_inc_input) or "SIN_CODIGO"
    downloads_dest = DOWNLOADS_ROOT / folder_name
    downloads_dest.mkdir(parents=True, exist_ok=True)
    log(f"📁 Carpeta de Descargas asegurada:\n{downloads_dest}")

    if inf_path and inf_path.is_file():
        try:
            dst = downloads_dest / inf_path.name
            shutil.copy2(inf_path, dst)
            log(f"📑 INF copiado a Descargas:\n{dst}")
        except Exception as e:
            log(f"❌ No se pudo copiar el INF a Descargas:\nOrigen: {inf_path}\nDestino: {downloads_dest}\nDetalle: {e}")
    else:
        log("ℹ️ No se copió INF a Descargas (no se generó o no existe el archivo).")

# ---------- Reemplazos dentro del Excel (ticket, título, fecha, desarrollador) ----------
PLACEHOLDER_TICKET     = "RITMXXXXXXX"
PLACEHOLDER_TITULO     = "INF-DDMMAAAA-RITMXXXXXXX-{SISTEMA}-{TITULO}"
PLACEHOLDER_FECHA_AAA  = "DD/MM/AAA"
PLACEHOLDER_FECHA_AAAA = "DD/MM/AAAA"
PLACEHOLDER_DEV        = "{especialista de soluciones}"

def _fecha_de_stem_ddmmaaaa_to_dd_mm_aaaa(stem: str) -> str | None:
    """Extrae DDMMAAAA del nombre del archivo (INF-DDMMAAAA-...) y lo devuelve 'DD/MM/AAAA'."""
    try:
        parts = stem.split("-")
        ddmmaaaa = parts[1]
        if len(ddmmaaaa) == 8 and ddmmaaaa.isdigit():
            return f"{ddmmaaaa[0:2]}/{ddmmaaaa[2:4]}/{ddmmaaaa[4:8]}"
    except Exception:
        pass
    return None

def replace_placeholders_in_workbook(xlsx_path: Path, ticket_code: str, title_text: str, dev_name: Optional[str]) -> bool:
    """
    Reemplaza:
      - RITMXXXXXXX -> ticket_code
      - INF-DDMMAAAA-RITMXXXXXXX-{SISTEMA}-{TITULO} -> title_text (stem del archivo)
      - DD/MM/AAA(A) -> fecha derivada del nombre del archivo
      - {especialista de soluciones} -> dev_name
    Si el placeholder del título no aparece, escribe title_text a la derecha de 'Título Informe'.
    Si el placeholder del desarrollador no aparece, escribe dev_name a la derecha de 'Analista Desarrollador'.
    """
    if not xlsx_path.is_file():
        log(f"⚠️  No existe el archivo para actualizar: {xlsx_path}")
        return False

    fecha_formateada = _fecha_de_stem_ddmmaaaa_to_dd_mm_aaaa(xlsx_path.stem)

    try:
        wb = load_workbook(filename=str(xlsx_path))
    except Exception as e:
        log(f"❌ No se pudo abrir el Excel para edición: {xlsx_path}\nDetalle: {e}")
        return False

    changed_any = False
    dev_written = False
    title_written = False

    for ws in wb.worksheets:
        # 1) Reemplazos directos en todo el texto de celdas
        for row in ws.iter_rows(values_only=False):
            for cell in row:
                val = cell.value
                if isinstance(val, str):
                    new_val = val

                    if PLACEHOLDER_TICKET in new_val:
                        new_val = new_val.replace(PLACEHOLDER_TICKET, ticket_code)

                    if PLACEHOLDER_TITULO in new_val:
                        new_val = new_val.replace(PLACEHOLDER_TITULO, title_text)
                        title_written = True

                    if fecha_formateada:
                        if PLACEHOLDER_FECHA_AAAA in new_val:
                            new_val = new_val.replace(PLACEHOLDER_FECHA_AAAA, fecha_formateada)
                        if PLACEHOLDER_FECHA_AAA in new_val:
                            new_val = new_val.replace(PLACEHOLDER_FECHA_AAA, fecha_formateada)

                    if dev_name and PLACEHOLDER_DEV in new_val:
                        new_val = new_val.replace(PLACEHOLDER_DEV, dev_name)
                        dev_written = True

                    if new_val != val:
                        cell.value = new_val
                        changed_any = True

        # 2) Si el título no se pudo por placeholder, escribir por etiqueta "Título Informe"
        if not title_written:
            for row in ws.iter_rows(values_only=False):
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.strip().lower() == "título informe":
                        try:
                            ws.cell(row=cell.row, column=cell.column + 1, value=title_text)
                            changed_any = True
                            title_written = True
                        except Exception:
                            pass
                        break
                if title_written:
                    break

        # 3) Si el desarrollador no se pudo por placeholder, escribir por etiqueta
        if dev_name and not dev_written:
            for row in ws.iter_rows(values_only=False):
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.strip().lower() == "analista desarrollador":
                        try:
                            ws.cell(row=cell.row, column=cell.column + 1, value=dev_name)
                            changed_any = True
                            dev_written = True
                        except Exception:
                            pass
                        break
                if dev_written:
                    break

    if changed_any:
        try:
            wb.save(str(xlsx_path))
            log(f"📝 Excel actualizado (ticket, título, fecha y desarrollador):\n{xlsx_path}")
            return True
        except Exception as e:
            log(f"❌ Error al guardar cambios en: {xlsx_path}\nDetalle: {e}")
            return False
    else:
        log("ℹ️ No se detectaron cambios para guardar en el Excel.")
        return False

# -------------------------------- MAIN --------------------------------
def main() -> None:
    # 1) Pide el RITM/INC
    ritm_inc_input = input("Ingresa el código RITM/INC (ej. RITM0557437 o INC1286211): ").strip()
    if not ritm_inc_input:
        log("No se ingresó un RITM/INC. Saliendo…")
        return

    # 2) Lee Excel
    if not EXCEL_PATH.exists():
        log(f"❌ No se encontró el archivo Excel en: {EXCEL_PATH}")
        return

    try:
        df = pd.read_excel(EXCEL_PATH, sheet_name=SHEET_NAME, engine="openpyxl")
    except Exception as e:
        log(f"❌ Error al leer el Excel: {e}")
        return

    # 3) Valida columnas (incluye DESARROLLADOR)
    required_cols = [COL_RITM_INC, COL_SCTASK, COL_APLICATIVO, COL_DESCRIPCION, COL_DESARROLLADOR]
    missing = [c for c in required_cols if c not in df.columns]
    if missing:
        log(f"❌ Faltan columnas en '{SHEET_NAME}': {missing}")
        return

    # 4) Busca fila
    row = find_ticket_row(df, ritm_inc_input)
    if row is None:
        log(f"❌ No se encontró el ticket '{ritm_inc_input}' en '{SHEET_NAME}'.")
        return

    # 5) Extrae campos
    sctask        = row[COL_SCTASK]
    aplicativo    = row[COL_APLICATIVO]
    descripcion   = row[COL_DESCRIPCION]
    desarrollador = row[COL_DESARROLLADOR] if COL_DESARROLLADOR in row.index else None

    # 6) Construye nombre de carpeta
    folder_name = build_folder_name(ritm_inc_input, sctask, aplicativo, descripcion)

    # 7) Crea carpeta destino
    DEST_ROOT.mkdir(parents=True, exist_ok=True)
    folder_path = DEST_ROOT / folder_name
    try:
        folder_path.mkdir(parents=True, exist_ok=False)
        log(f"Carpeta creada:\n{folder_path}")
    except FileExistsError:
        log(f"ℹ️ La carpeta ya existía:\n{folder_path}")
    except OSError as e:
        log(f"❌ Error al crear la carpeta:\n{folder_path}\nDetalle: {e}")
        return

    # 8) Copia estructura de plantilla
    ok_template = copy_template_into(folder_path)
    if not ok_template:
        log("⚠️  No se copió la estructura de plantilla (verifica ruta/OneDrive).")

    # 9) Copia y renombra INF en '3. Informe Certificación'
    inf_path = copy_and_rename_inf(folder_path, folder_name)

    # 9.1) Si se generó el INF, actualiza placeholders internos (ticket, título, fecha, dev)
    if inf_path:
        ticket_code = ritm_inc_input.strip()
        title_text  = inf_path.stem
        dev_name    = None if pd.isna(desarrollador) else str(desarrollador).strip()
        replace_placeholders_in_workbook(inf_path, ticket_code, title_text, dev_name)

    # 10) Crea SIEMPRE la carpeta en Descargas y, si hay INF, lo copia
    ensure_downloads_folder_and_copy(ritm_inc_input, inf_path)

    # 11) Resumen
    log("\nResumen:")
    log(f"  🗂  Carpeta destino: {folder_path}")
    log(f"  📦 Plantilla copiada: {'Sí' if ok_template else 'No'}")
    log(f"  📑 INF generado: {'Sí' if inf_path else 'No'}")
    log(f"  ⬇️  Carpeta en Descargas: {DOWNLOADS_ROOT / (sanitize_segment(ritm_inc_input) or 'SIN_CODIGO')}")

if __name__ == "__main__":
    main()
