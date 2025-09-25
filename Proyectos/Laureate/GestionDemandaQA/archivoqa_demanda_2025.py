import os
import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, NamedStyle

# Rutas de carpetas a analizar
folder_paths = [
    r"C:\Users\manue\OneDrive - EduCorpPERU\2025\Tickets",
    r"C:\Users\manue\OneDrive - EduCorpPERU\2025\Tickets\Finalizados\2025"
]

# Ruta donde se exportará el archivo Excel
export_path = os.path.join(folder_paths[0], "Exported_Tickets.xlsx")

# Ruta de los archivos adicionales
certificaciones_path = r"C:\Users\manue\OneDrive - EduCorpPERU\Proyectos QA - 2025\Gestión Demanda Certificaciones.xlsx"
change_request_path = os.path.join(folder_paths[0], "change_request.xlsx")

# Leer los archivos Excel
certificaciones_df = pd.read_excel(certificaciones_path, sheet_name="Tickets Diarios")
tickets_df = pd.read_excel(change_request_path, sheet_name="Page 1")

# Cargar el archivo de change_request.xlsx para obtener State, Created y Updated
df_change_request = pd.read_excel(change_request_path, sheet_name="Page 1")

# Verificar los nombres de las columnas disponibles
print("Columnas disponibles en change_request.xlsx:", df_change_request.columns)

# Normalizar nombres de columnas
df_change_request.columns = df_change_request.columns.str.strip()

# Verificar que las columnas necesarias existen
required_columns = ["Number", "State", "Created", "Updated"]
missing_columns = [col for col in required_columns if col not in df_change_request.columns]
if missing_columns:
    raise KeyError(f"Las siguientes columnas faltan en change_request.xlsx: {missing_columns}")

# Asegurar que las columnas "Created" y "Updated" sean de tipo datetime
df_change_request["Created"] = pd.to_datetime(df_change_request["Created"], errors='coerce')
df_change_request["Updated"] = pd.to_datetime(df_change_request["Updated"], errors='coerce')

chg_info = df_change_request.set_index("Number")[["State", "Created", "Updated"]].to_dict("index")

# Función para extraer el tercer segmento del nombre de la carpeta
def extract_third_segment(folder_name):
    parts = folder_name.split("-")
    return parts[2] if len(parts) >= 3 else ""

# Lista para almacenar los datos
data = []

# Recorrer ambas carpetas
for folder_path in folder_paths:
    for folder_name in os.listdir(folder_path):
        full_path = os.path.join(folder_path, folder_name)

        if os.path.isdir(full_path):
            # Buscar patrones RITM/INC y SCTASK
            ritm_inc_match = re.search(r'(RITM\d+|INC\d+)', folder_name)
            sctask_match = re.search(r'(SCTASK\d+)', folder_name)

            ritm_inc = ritm_inc_match.group(0) if ritm_inc_match else ""
            sctask = sctask_match.group(0) if sctask_match else ""
            third_segment = extract_third_segment(folder_name)

            # Obtener CHG asociado
            chg = ""
            search_value = ritm_inc if ritm_inc else third_segment
            if search_value:
                chg_values = tickets_df.loc[tickets_df["Short description"].str.contains(search_value, na=False, case=False, regex=False), "Number"].values
                chg = chg_values[0] if len(chg_values) > 0 else ""

            # Obtener detalles del CHG
            chg_data = chg_info.get(chg, {"State": "", "Created": None, "Updated": None})
            state, created, updated = chg_data.get("State", ""), chg_data.get("Created", None), chg_data.get("Updated", None)

            # Convertir valores de fecha a tipo fecha sin hora
            created = created.date() if isinstance(created, pd.Timestamp) else None
            updated = updated.date() if isinstance(updated, pd.Timestamp) else None

            # Dividir el nombre de la carpeta por guiones
            fragments = folder_name.split("-")
            portal = fragments[4] if len(fragments) > 4 else ""
            descripcion = fragments[5] if len(fragments) > 5 else ""

            # Buscar "NUM." en la hoja "Tickets Diarios"
            num = certificaciones_df.loc[certificaciones_df["RITM/INC"] == ritm_inc, "NUM."].values
            num = num[0] if len(num) > 0 else ""

            # Determinar la fuente (Tickets o Finalizados)
            fuente = "Finalizados" if "Finalizados" in folder_path else "Tickets"

            # Agregar a la lista
            if ritm_inc or sctask or descripcion or portal:
                data.append({
                    "NUM.": num,
                    "RITM/INC": ritm_inc, 
                    "SCTASK": sctask,                 
                    "DESCRIPCIÓN": descripcion, 
                    "PORTAL": portal,
                    "CHG": chg,
                    "Fuente": fuente,
                    "State": state,
                    "Created": created,
                    "Updated": updated
                })

# Crear DataFrame
df = pd.DataFrame(data)

# Exportar a Excel con formato de fecha
df.to_excel(export_path, index=False)

# **Aplicar formato de color a "Finalizados" y establecer formato de fecha**
wb = load_workbook(export_path)
ws = wb.active

# Crear estilo de fecha
date_style = NamedStyle(name="date_style", number_format="YYYY-MM-DD")

# Aplicar formato de fecha a las columnas "Created" y "Updated"
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=9, max_col=10):  # Columnas I y J
    for cell in row:
        cell.style = date_style

# Definir color para celdas de la fuente "Finalizados"
fill_finalizados = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")

# Aplicar color a filas de "Finalizados"
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    if row[6].value == "Finalizados":  # Columna "Fuente"
        for cell in row:
            cell.fill = fill_finalizados

# Guardar cambios
wb.save(export_path)

print(f"Archivo Excel exportado correctamente en: {export_path}")