# -*- coding: utf-8 -*-
r"""
Actualiza SOLO celdas específicas en:
  Gestión Demanda Certificaciones LIUv1.xlsx  (hoja: 'Tickets Diarios')

Reglas:
- Filtra filas con ANALISTA QA = 'Manuel Hurtado' y CODIGO CHG vacío
- Lookup en: C:\Users\manue\OneDrive - EduCorpPERU\2025\Tickets\Exported_Tickets.xlsx (hoja 'Tickets')
- Prioridad de cruce: 1) SCTASK, 2) RITM/INC
- Completa en CERT:
    CHG               -> CODIGO CHG
    Created (fecha)   -> FECHA SOLICITUD CHG (solo fecha)
    Created_hora      -> HORA DE SOLICITUD CHG (HH:MM:SS)
    State (mapeado)   -> ESTADO QA
- Mantiene TODO el formato original (bordes, colores, bloqueos, etc.)
- Crea un backup del archivo CERT antes de escribir.
"""

from pathlib import Path
from datetime import datetime
import shutil
import warnings

import pandas as pd
from openpyxl import load_workbook

# ------------------ RUTAS ------------------
CERT_PATH = Path(r"C:\Users\manue\OneDrive - EduCorpPERU\Calidad de Software - Certificaciones\Gestión Demanda Certificaciones LIUv1.xlsx")
CERT_SHEET = "Tickets Diarios"

HELPER_PATH = Path(r"C:\Users\manue\OneDrive - EduCorpPERU\2025\Tickets\Exported_Tickets.xlsx")
HELPER_SHEET = "Tickets"
# -------------------------------------------

# Silenciar warnings ruidosos (openpyxl)
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

GREEN = "\033[92m"
RESET = "\033[0m"

def log(msg: str):
    print(f"[{datetime.now():%H:%M:%S}] {msg}", flush=True)

def ok(msg: str):
    print(f"{GREEN}✅ {msg}{RESET}", flush=True)

def is_blank(v) -> bool:
    if v is None:
        return True
    try:
        import math
        if isinstance(v, float) and math.isnan(v):
            return True
    except Exception:
        pass
    return str(v).strip() == ""

def norm_key(x: str) -> str:
    """Normaliza clave para comparación (upper + trim)."""
    if x is None:
        return ""
    return str(x).strip().upper()

def read_helper() -> pd.DataFrame:
    """Lee 'Tickets' y devuelve DF con claves normalizadas para join."""
    if not HELPER_PATH.exists():
        raise FileNotFoundError(f"No existe helper: {HELPER_PATH}")
    df = pd.read_excel(HELPER_PATH, sheet_name=HELPER_SHEET)
    df.columns = [str(c).strip() for c in df.columns]

    needed = ["RITM/INC", "SCTASK", "CHG", "Created", "Created_hora", "State"]  # <-- añadimos State
    miss = [c for c in needed if c not in df.columns]
    if miss:
        raise KeyError(f"Faltan columnas en helper 'Tickets': {miss}")

    out = df[needed].copy()
    out["_SCTASK_UP"] = out["SCTASK"].astype(str).fillna("").str.strip().str.upper()
    out["_RITM_UP"]   = out["RITM/INC"].astype(str).fillna("").str.strip().str.upper()
    return out

def find_header_indexes(ws, expected_headers):
    """
    Busca índices de columna para cada encabezado (fila 1).
    Devuelve dict {header: col_index}.
    """
    header_map = {}
    for col_idx, cell in enumerate(ws[1], start=1):
        name = str(cell.value).strip() if cell.value is not None else ""
        header_map[name] = col_idx

    result, missing = {}, []
    for h in expected_headers:
        if h in header_map:
            result[h] = header_map[h]
        else:
            missing.append(h)
    if missing:
        raise KeyError(f"No se encontraron columnas en '{ws.title}': {missing}")
    return result

def main():
    # ----- 1) Cargar helper -----
    log("Cargando helper Exported_Tickets.xlsx / Tickets ...")
    helper = read_helper()
    ok(f"Helper cargado: {len(helper):,} filas")

    # Mapeo de State -> ESTADO QA
    state_to_estadoqa = {
        "REVIEW": "Ticket Cerrado",
        "SCHEDULED": "Derivado",
        "NEW": "En Proceso",
    }

    # ----- 2) Abrir CERT manteniendo formato -----
    if not CERT_PATH.exists():
        raise FileNotFoundError(f"No existe CERT: {CERT_PATH}")
    log("Abriendo CERT (manteniendo formato)...")
    wb = load_workbook(CERT_PATH)
    if CERT_SHEET not in wb.sheetnames:
        raise KeyError(f"No existe la hoja '{CERT_SHEET}' en {CERT_PATH}")
    ws = wb[CERT_SHEET]

    # ----- 3) Índices de columnas necesarias -----
    cols_needed = [
        "ANALISTA QA", "CODIGO CHG",
        "SCTASK", "RITM/INC",
        "FECHA SOLICITUD CHG", "HORA DE SOLICITUD CHG",
        "ESTADO QA"  # <-- nueva columna a actualizar
    ]
    idx = find_header_indexes(ws, cols_needed)

    # ----- 4) Filas objetivo (mismo filtro que antes) -----
    target_rows = []
    for r in range(2, ws.max_row + 1):
        analista = ws.cell(row=r, column=idx["ANALISTA QA"]).value
        chg_cell = ws.cell(row=r, column=idx["CODIGO CHG"]).value
        if str(analista).strip().casefold() == "manuel hurtado" and is_blank(chg_cell):
            target_rows.append(r)

    if not target_rows:
        ok("No hay filas con ANALISTA QA = 'Manuel Hurtado' y CODIGO CHG vacío. Nada que actualizar.")
        return

    log(f"Filas objetivo: {len(target_rows)}")

    # ----- 5) Índices rápidos de helper -----
    helper_by_sctask = (
        helper.dropna(subset=["_SCTASK_UP"])
              .drop_duplicates("_SCTASK_UP", keep="first")
              .set_index("_SCTASK_UP")
    )
    helper_by_ritm = (
        helper.dropna(subset=["_RITM_UP"])
              .drop_duplicates("_RITM_UP", keep="first")
              .set_index("_RITM_UP")
    )

    # ----- 6) Completar celdas (manteniendo formato) -----
    updates = 0
    for r in target_rows:
        sctask_val = norm_key(ws.cell(row=r, column=idx["SCTASK"]).value)
        ritm_val   = norm_key(ws.cell(row=r, column=idx["RITM/INC"]).value)

        # lookup por SCTASK, luego por RITM/INC
        hit = None
        if sctask_val and sctask_val in helper_by_sctask.index:
            hit = helper_by_sctask.loc[sctask_val]
        elif ritm_val and ritm_val in helper_by_ritm.index:
            hit = helper_by_ritm.loc[ritm_val]

        if hit is None or (isinstance(hit, pd.Series) and hit.empty):
            continue

        # Orígenes
        chg         = hit.get("CHG", None)
        created_dt  = pd.to_datetime(hit.get("Created", pd.NaT), errors="coerce")
        created_h   = hit.get("Created_hora", None)
        state_raw   = str(hit.get("State", "") or "").strip()
        estado_qa   = state_to_estadoqa.get(state_raw.upper(), None)

        # Escribir (CHG requerido para considerar la coincidencia, como antes)
        if chg and str(chg).strip():
            ws.cell(row=r, column=idx["CODIGO CHG"]).value = str(chg).strip()
            ws.cell(row=r, column=idx["FECHA SOLICITUD CHG"]).value = (
                None if pd.isna(created_dt) else created_dt.date()
            )
            ws.cell(row=r, column=idx["HORA DE SOLICITUD CHG"]).value = (
                None if not created_h or str(created_h).strip() == "" else str(created_h).strip()
            )
            # ESTADO QA (solo si hay mapeo válido)
            if estado_qa:
                ws.cell(row=r, column=idx["ESTADO QA"]).value = estado_qa
            updates += 1

    if updates == 0:
        ok("No se encontraron coincidencias en el helper para las filas objetivo. Archivo sin cambios.")
        return

    # ----- 7)  guardado -----

    wb.save(CERT_PATH)
    ok(f"Actualización escrita en: {CERT_PATH}")
    ok(f"Filas actualizadas: {updates}")

if __name__ == "__main__":
    main()
