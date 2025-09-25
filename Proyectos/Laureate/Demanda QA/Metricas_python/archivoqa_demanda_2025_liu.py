# -*- coding: utf-8 -*-
r"""
Exporta tres hojas:
- "Tickets": carpetas activas con nueva estructura (sin NUM.)
- "Finalizados": carpetas finalizadas con lógica anterior (con NUM.)
- "Métricas": métricas obtenidas EXCLUSIVAMENTE de la hoja "Tickets Diarios" (CERT_PATH)

Métrica actual:
- Conteo de registros por "ANALISTA QA" (usa 'SIN ASIGNAR' cuando está vacío)

Requisitos: pandas, openpyxl
"""

import os
import re
import sys
from datetime import datetime
from pathlib import Path
from typing import Dict, Any, Iterable, Tuple

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, NamedStyle
from openpyxl.utils import get_column_letter


# =============== CONFIG ===============
FOLDER_TICKETS = Path(r"C:\Users\manue\OneDrive - EduCorpPERU\2025\Tickets")
FOLDER_FINALIZADOS = Path(r"C:\Users\manue\OneDrive - EduCorpPERU\2025\Tickets\Finalizados\2025")

CHANGE_REQUEST_PATH = FOLDER_TICKETS / "change_request.xlsx"
CHANGE_REQUEST_SHEET = "Page 1"   # cambia si tu export usa otro nombre de hoja

# CERT: TODAS LAS MÉTRICAS SALEN DE AQUÍ
CERT_PATH = Path(r"C:\Users\manue\OneDrive - EduCorpPERU\Calidad de Software - Certificaciones\Gestión Demanda Certificaciones LIUv1.xlsx")
CERT_SHEET = "Tickets Diarios"

EXPORT_PATH = FOLDER_TICKETS / f"Exported_Tickets_{datetime.now():%Y%m%d_%H%M%S}.xlsx"

# Columnas finales históricas (para "Finalizados")
TARGET_COLS_WITH_NUM = [
    "NUM.", "RITM/INC", "SCTASK", "DESCRIPCIÓN", "PORTAL",
    "CHG", "Fuente", "State", "Created", "Updated"
]
# Para "Tickets" (sin NUM.)
TARGET_COLS_NO_NUM = [
    "RITM/INC", "SCTASK", "PORTAL", "DESCRIPCIÓN",
    "CHG", "Fuente", "State", "Created", "Updated"
]
# ======================================


# ------------- Utilidades -------------
def log(msg: str):
    print(f"[{datetime.now():%H:%M:%S}] {msg}")


def _normalize_headers(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.columns = [str(c).strip() for c in df.columns]
    return df


def _to_datetime(df: pd.DataFrame, cols):
    for c in cols:
        if c in df.columns:
            df[c] = pd.to_datetime(df[c], errors="coerce")
    return df


def safe_str(x) -> str:
    return "" if pd.isna(x) else str(x).strip()


def safe_read_excel(path: Path, sheet_name: str | None = None) -> pd.DataFrame:
    if not path.exists():
        raise FileNotFoundError(f"No se encontró el archivo: {path}")
    return pd.read_excel(path, sheet_name=sheet_name)
# --------------------------------------


# --------- Fuentes de datos -----------
def load_change_request(path: Path, sheet: str) -> pd.DataFrame:
    df = safe_read_excel(path, sheet)
    df = _normalize_headers(df)
    # alias
    rmap = {}
    for c in df.columns:
        lc = c.strip().lower()
        if lc == "número": rmap[c] = "Number"
        elif lc == "descripción breve": rmap[c] = "Short description"
    if rmap:
        df = df.rename(columns=rmap)

    must = {"Number", "Short description"}
    if not must <= set(df.columns):
        raise KeyError(f"change_request debe incluir {must}. Columnas: {list(df.columns)}")

    df = _to_datetime(df, ["Created", "Updated"])
    keep = [c for c in ["Number", "Short description", "State", "Created", "Updated"] if c in df.columns]
    return df[keep].copy()


def load_tickets_diarios(path: Path, sheet: str) -> pd.DataFrame:
    df = safe_read_excel(path, sheet)
    return _normalize_headers(df)
# --------------------------------------


# ----------- Parsing carpetas ----------
def parse_for_finalizados(folder_name: str):
    """
    Lógica ACTUAL:
      portal = segmento [4] (5.º)
      desc   = segmento [5] (6.º)
      ritm/inc y sctask por regex
      third  = segmento [2] (3.º)
    """
    ritm_inc_match = re.search(r"(RITM\d+|INC\d+)", folder_name, flags=re.IGNORECASE)
    sctask_match = re.search(r"(SCTASK\d+)", folder_name, flags=re.IGNORECASE)
    parts = folder_name.split("-")
    ritm_inc = ritm_inc_match.group(0).upper() if ritm_inc_match else ""
    sctask = sctask_match.group(0).upper() if sctask_match else ""
    portal = parts[4] if len(parts) > 4 else ""
    descripcion = parts[5] if len(parts) > 5 else ""
    third_segment = parts[2] if len(parts) > 2 else ""
    return ritm_inc, sctask, portal, descripcion, third_segment


def parse_for_tickets(folder_name: str):
    """
    Estructura para 'Tickets':
      [0]=Fecha, [1]=RITM/INC, [2]=SCTASK (u otro), [3]=Portal, [4]=Descripción, [*]=extras
      - RITM/INC: segundo segmento si coincide; si no, regex
      - SCTASK: regex en todo el nombre
    """
    import re

    parts = folder_name.split("-")

    # RITM/INC
    ritm_inc = ""
    if len(parts) > 1 and re.match(r"^(RITM\d+|INC\d+)$", parts[1], flags=re.IGNORECASE):
        ritm_inc = parts[1].upper()
    else:
        m = re.search(r"(RITM\d+|INC\d+)", folder_name, flags=re.IGNORECASE)
        ritm_inc = m.group(0).upper() if m else ""

    # SCTASK
    sctask_match = re.search(r"(SCTASK\d+)", folder_name, flags=re.IGNORECASE)
    sctask = sctask_match.group(0).upper() if sctask_match else ""

    # 4.º segmento -> PORTAL ; 5.º -> DESCRIPCIÓN
    portal = parts[3] if len(parts) > 3 else ""
    descripcion = parts[4] if len(parts) > 4 else ""

    # Devuelve también el 3er segmento por compatibilidad ascendente (si lo usas en otra parte)
    third_segment = parts[2] if len(parts) > 2 else ""

    return ritm_inc, sctask, portal, descripcion, third_segment

def find_chg_by_token(change_df: pd.DataFrame, token: str) -> str:
    token = safe_str(token)
    if not token or change_df.empty:
        return ""
    mask = change_df["Short description"].astype(str).str.contains(token, case=False, regex=False, na=False)
    subset = change_df.loc[mask].copy()
    if subset.empty:
        return ""
    subset["Number"] = subset["Number"].astype(str)
    subset = subset[subset["Number"].str.upper().str.startswith("CHG")]
    if subset.empty:
        return ""
    if "Updated" in subset.columns:
        subset = subset.sort_values("Updated", ascending=False)
    return safe_str(subset.iloc[0]["Number"])


def enrich_chg_meta(change_df: pd.DataFrame, chg: str) -> Dict[str, Any]:
    out = {"CHG": chg or "", "State": pd.NA, "Created": pd.NA, "Updated": pd.NA}
    if not chg or change_df.empty:
        return out
    row = change_df.loc[change_df["Number"].astype(str).str.upper() == chg.upper()]
    if row.empty:
        return out
    r = row.iloc[0]
    for k in ["State", "Created", "Updated"]:
        if k in r.index:
            out[k] = r[k]
    return out
# --------------------------------------


# ------- NUM. desde Tickets Diarios ----
def lookup_num(cert_df: pd.DataFrame, sctask: str, ritm_inc: str) -> str:
    if "NUM." not in cert_df.columns:
        return ""
    if "SCTASK" in cert_df.columns and sctask:
        vals = cert_df.loc[cert_df["SCTASK"].astype(str).str.upper() == sctask.upper(), "NUM."].values
        if len(vals) > 0:
            return safe_str(vals[0])
    if "RITM/INC" in cert_df.columns and ritm_inc:
        vals = cert_df.loc[cert_df["RITM/INC"].astype(str).str.upper() == ritm_inc.upper(), "NUM."].values
        if len(vals) > 0:
            return safe_str(vals[0])
    return ""
# --------------------------------------


# -------------- Excel estilos ---------------
def autosize(ws):
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in col_cells:
            val = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(val))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)


def style_dates(ws):
    date_style = NamedStyle(name="date_style", number_format="yyyy-mm-dd")
    header = [c.value for c in ws[1]]
    for dc in ("Created", "Updated"):
        if dc in header:
            j = header.index(dc) + 1
            for r in range(2, ws.max_row + 1):
                cell = ws.cell(row=r, column=j)
                if isinstance(cell.value, datetime):
                    cell.style = date_style


def highlight_finalizados(ws):
    header = [c.value for c in ws[1]]
    if "Fuente" not in header:
        return
    idx = header.index("Fuente") + 1
    fill = PatternFill(start_color="FFF8B3", end_color="FFF8B3", fill_type="solid")
    for r in range(2, ws.max_row + 1):
        if (ws.cell(row=r, column=idx).value or "").strip() == "Finalizados":
            for c in range(1, ws.max_column + 1):
                ws.cell(row=r, column=c).fill = fill
# -------------------------------------------


# -------- Construcción de cada hoja --------
def build_sheet_rows(base_folder: Path,
                     fuente: str,
                     parser,          # parse_for_tickets o parse_for_finalizados
                     change_df: pd.DataFrame,
                     cert_df: pd.DataFrame,
                     drop_num: bool) -> pd.DataFrame:
    rows = []
    if not base_folder.exists():
        log(f"[WARN] No existe la carpeta: {base_folder}")
        return pd.DataFrame(columns=TARGET_COLS_NO_NUM if drop_num else TARGET_COLS_WITH_NUM)

    for entry in os.scandir(base_folder):
        if not entry.is_dir():
            continue
        folder_name = entry.name
        if folder_name.startswith("~") or folder_name.startswith("."):
            continue

        ritm_inc, sctask, portal, descripcion, third_segment = parser(folder_name)

        # CHG: SCTASK -> RITM/INC -> tercer segmento
        chg = ""
        if sctask:
            chg = find_chg_by_token(change_df, sctask)
        if not chg and ritm_inc:
            chg = find_chg_by_token(change_df, ritm_inc)
        if not chg and third_segment:
            chg = find_chg_by_token(change_df, third_segment)

        meta = enrich_chg_meta(change_df, chg)

        row = {
            "RITM/INC": ritm_inc,
            "SCTASK": sctask,
            "DESCRIPCIÓN": descripcion,
            "PORTAL": portal,
            "CHG": meta.get("CHG", ""),
            "Fuente": fuente,
            "State": meta.get("State", pd.NA),
            "Created": meta.get("Created", pd.NA),
            "Updated": meta.get("Updated", pd.NA),
        }

        if not drop_num:
            row["NUM."] = lookup_num(cert_df, sctask, ritm_inc)

        if ritm_inc or sctask or descripcion or portal:
            rows.append(row)

    df = pd.DataFrame(rows)
    target = TARGET_COLS_NO_NUM if drop_num else TARGET_COLS_WITH_NUM
    for c in target:
        if c not in df.columns:
            df[c] = pd.NA
    df = df[target].copy()

    for dc in ["Created", "Updated"]:
        df[dc] = pd.to_datetime(df[dc], errors="coerce")

    subset_cols = ["RITM/INC", "SCTASK", "DESCRIPCIÓN", "PORTAL", "CHG"]
    df = df.drop_duplicates(subset=subset_cols, keep="first")

    return df
# -------------------------------------------


# ----------- MÉTRICAS (SOLO CERT) ----------
def build_metrics_from_cert(df_cert: pd.DataFrame) -> pd.DataFrame:
    """
    Construye métricas SOLO desde CERT_PATH/hoja 'Tickets Diarios'.
    Métrica actual: Conteo por ANALISTA QA.
    """
    if "ANALISTA QA" not in df_cert.columns:
        raise KeyError("No se encontró la columna 'ANALISTA QA' en 'Tickets Diarios'.")

    s = (df_cert["ANALISTA QA"]
         .fillna("")
         .astype(str)
         .str.strip())
    s = s.replace({"": "SIN ASIGNAR"})

    metrics = (s.value_counts()
                 .rename("Conteo Total")
                 .reset_index()
                 .rename(columns={"index": "ANALISTA QA"})
                 .sort_values(["Conteo Total", "ANALISTA QA"], ascending=[False, True])
              )
    return metrics[["ANALISTA QA", "Conteo Total"]]
# -------------------------------------------


# ----------------- Main --------------------
def main():
    try:
        log("Cargando change_request.xlsx ...")
        df_change = load_change_request(CHANGE_REQUEST_PATH, CHANGE_REQUEST_SHEET)
        log(f"change_request: {len(df_change):,} filas")

        log("Cargando 'Tickets Diarios' (CERT_PATH) ...")
        df_cert = load_tickets_diarios(CERT_PATH, CERT_SHEET)
        log(f"Tickets Diarios: {len(df_cert):,} filas")

        # Hoja 'Tickets' (nueva estructura, sin NUM.)
        log("Construyendo hoja 'Tickets' ...")
        df_tickets = build_sheet_rows(
            base_folder=FOLDER_TICKETS,
            fuente="Tickets",
            parser=parse_for_tickets,
            change_df=df_change,
            cert_df=df_cert,
            drop_num=True
        )
        log(f"'Tickets': {len(df_tickets):,} filas x {len(df_tickets.columns)} cols")

        # Hoja 'Finalizados' (lógica actual, con NUM.)
        log("Construyendo hoja 'Finalizados' ...")
        df_final = build_sheet_rows(
            base_folder=FOLDER_FINALIZADOS,
            fuente="Finalizados",
            parser=parse_for_finalizados,
            change_df=df_change,
            cert_df=df_cert,
            drop_num=False
        )
        log(f"'Finalizados': {len(df_final):,} filas x {len(df_final.columns)} cols")

        # Hoja 'Métricas' (EXCLUSIVO de CERT_PATH)
        log("Construyendo hoja 'Métricas' (desde CERT_PATH) ...")
        df_metrics = build_metrics_from_cert(df_cert)
        log(f"'Métricas': {len(df_metrics):,} filas x {len(df_metrics.columns)} cols")

        # Exportar
        log(f"Exportando a: {EXPORT_PATH}")
        with pd.ExcelWriter(EXPORT_PATH, engine="openpyxl") as writer:
            df_tickets.to_excel(writer, sheet_name="Tickets", index=False)
            df_final.to_excel(writer, sheet_name="Finalizados", index=False)
            df_metrics.to_excel(writer, sheet_name="Métricas", index=False)

        wb = load_workbook(EXPORT_PATH)

        # Formato por hoja
        for sheet_name in ["Tickets", "Finalizados", "Métricas"]:
            ws = wb[sheet_name]
            if sheet_name in ["Tickets", "Finalizados"]:
                style_dates(ws)
                if sheet_name == "Finalizados":
                    highlight_finalizados(ws)
            autosize(ws)
            ws.auto_filter.ref = ws.dimensions
            ws.freeze_panes = "A2"

        wb.save(EXPORT_PATH)
        log("OK. Listo.")

    except Exception as e:
        log(f"[ERROR] {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()
