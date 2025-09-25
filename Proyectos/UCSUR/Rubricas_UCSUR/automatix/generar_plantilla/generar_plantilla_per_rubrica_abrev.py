
import os
import re
import pandas as pd
from docx import Document
from openpyxl import load_workbook

# ====== CONFIGURAR LA RUTA PRINCIPAL ======
base_folder_path = r"C:\Users\manue\Downloads\Medicina Veterinaria y Zootecnia - Copy\Salud Pública Veterinaria y Zoonosis"
error_log = []
validation_log = []
document_log = []

ABREVIATURAS_REGEX = r"(EC1|EC2|EC3|EC4|EF|EP|ED)"

def extraer_abreviatura(nombre_archivo):
    match = re.search(ABREVIATURAS_REGEX, nombre_archivo)
    return match.group(1) if match else None

def extraer_puntaje_desde_texto(texto):
    patrones = [
        r"\((\d+)\s*(?:puntos?|pts?|pto\.?)\)",  # (4 puntos)
        r"(\d+)\s*(?:puntos?|pts?|pto\.?)"         # 4 puntos
    ]
    for patron in patrones:
        match = re.search(patron, texto, flags=re.IGNORECASE)
        if match:
            return int(match.group(1))
    return 0

def procesar_documento(word_file_path):
    doc = Document(word_file_path)
    tables = doc.tables
    evaluacion_data = []

    if not tables:
        error_log.append([word_file_path, "El documento no contiene tablas procesables"])
        return None

    for table_index, table in enumerate(tables):
        rows = table.rows
        if len(rows) < 2:
            continue

        for i in range(1, len(rows), 2):
            if i + 1 >= len(rows):
                continue

            desc_row = rows[i]
            puntaje_row = rows[i + 1]

            if len(desc_row.cells) < 4 or len(puntaje_row.cells) < 4:
                continue

            criterio = desc_row.cells[0].text.strip()
            desc_ld = desc_row.cells[1].text.strip()
            desc_l = desc_row.cells[2].text.strip()
            desc_nl = desc_row.cells[3].text.strip()

            if not criterio or not desc_ld or not desc_l or not desc_nl:
                continue

            # Intentar extraer de la fila de puntaje
            p_ld = extraer_puntaje_desde_texto(puntaje_row.cells[1].text.strip())
            p_l = extraer_puntaje_desde_texto(puntaje_row.cells[2].text.strip())
            p_nl = extraer_puntaje_desde_texto(puntaje_row.cells[3].text.strip())

            # Si los puntajes no están en la fila siguiente, buscar embebidos
            if p_ld == 0: p_ld = extraer_puntaje_desde_texto(desc_ld)
            if p_l == 0: p_l = extraer_puntaje_desde_texto(desc_l)
            if p_nl == 0: p_nl = extraer_puntaje_desde_texto(desc_nl)

            evaluacion_data.append([
                criterio,
                p_ld, "Logro Destacado", desc_ld,
                p_l, "Logrado", desc_l,
                p_nl, "No Logrado", desc_nl
            ])

    if not evaluacion_data:
        error_log.append([word_file_path, "El documento no generó datos válidos"])
        return None

    return pd.DataFrame(evaluacion_data, columns=[
        "ASPECTO / CRITERIO A EVALUAR", "Puntaje de la calificación", "Título de la calificación", 
        "Descripción de la calificación .1", "Puntaje de la calificación.1", "Título de la calificación .1", 
        "Descripción de la calificación .1", "Puntaje de la calificación.2", "Título de la calificación .2", 
        "Descripción de la calificación .2"
    ])

for root, dirs, files in os.walk(base_folder_path):
    word_files = [f for f in files if f.endswith(".docx")]

    if word_files:
        folder_name = os.path.basename(root)
        output_excel_path = os.path.join(root, f"{folder_name}.xlsx")

        excel_writer = pd.ExcelWriter(output_excel_path, engine="openpyxl")
        generated_sheets = []

        for filename in word_files:
            word_file_path = os.path.join(root, filename)
            df_reporte = procesar_documento(word_file_path)
            abreviatura = extraer_abreviatura(filename)

            if abreviatura is None:
                error_log.append([word_file_path, "No se encontró una abreviatura válida en el nombre del archivo"])
                document_log.append([filename, "Ignorado (sin abreviatura válida)"])
                continue

            sheet_name = abreviatura
            if sheet_name in generated_sheets:
                counter = 1
                while f"{sheet_name}_{counter}" in generated_sheets:
                    counter += 1
                sheet_name = f"{sheet_name}_{counter}"

            if df_reporte is not None:
                df_reporte.to_excel(excel_writer, sheet_name=sheet_name, index=False)
                generated_sheets.append(sheet_name)

            document_log.append([filename, "Procesado" if df_reporte is not None else "Ignorado"])

        excel_writer.close()
        print(f"✅ Reporte generado en: {output_excel_path}")

        wb = load_workbook(output_excel_path)
        if len(wb.sheetnames) != len(word_files):
            validation_log.append([output_excel_path, len(word_files), len(wb.sheetnames), "Diferencia detectada"])

if error_log:
    pd.DataFrame(error_log, columns=["Archivo", "Descripción del Problema"]).to_excel(
        os.path.join(base_folder_path, "Reporte_Errores.xlsx"), index=False)
    print("⚠️ Se generó un reporte de errores.")

if validation_log:
    pd.DataFrame(validation_log, columns=["Archivo Excel", "Archivos Word", "Hojas en Excel", "Observación"]).to_excel(
        os.path.join(base_folder_path, "Reporte_Validaciones.xlsx"), index=False)
    print("⚠️ Se generó un reporte de validaciones.")

#pd.DataFrame(document_log, columns=["Documento Word", "Estado"]).to_excel(
    #os.path.join(base_folder_path, "Reporte_Documentos_Procesados.xlsx"), index=False)
#print("📋 Se generó un reporte de documentos procesados.")
